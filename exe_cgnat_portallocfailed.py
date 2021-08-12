# Purpose: CGNAT Port Allocation Failed info gathering
# Author: Frank Chen
# Logics:
# 1. read so ips from files
# 2. for each ip:
#    A. ssh to the ip
#    B. 取得檔案清單(程式執行當日) = log_files:
#       shell ls -lt /var/log/ns.log*
#       以today篩選清單
#    C. 針對每一個log_file(device端):
#       a. ssh以grep取得過濾資訊(shell grep -i "portall" log_file)，若為gz檔案則使用zgrep
#       b. 針對ssh的輸出(std_out)，需考慮無資料時的處理方式(防呆)
#          去除不必要資訊(如warning、done)
#          接輸出儲存至相應log檔案(去除.gz副檔名)
#    D. 針對每個log_file(應用程式端):
#       a. count total lines
#       b. count occurrences of each unique lsnSubscrIP
#       c. 輸出 a 到xlsx的so頁簽
#       d. 輸出 b 的top 10到so頁簽
#  ** 目錄結構: app_path/xtime/so_name_ip.txt
#  **                  /cgnat_portall_xtime.xlsx(sheet_name=so)


import datetime
import os
import sys

import openpyxl
import paramiko
from datetime import date

debug_mode = False
work_dir = os.getcwd()
device_user = "nsroot"
device_pw = "nsroot"
ip_sheet_name = "IP"
# op_ips = []


class PMIPCOLS():
    col_ip = 1
    col_host = 2
    col_so = 3
    col_device_type = 4
    col_user = 5
    col_pw = 6


def xlsx_get_sites_and_telnet_infos(ip_xlsx_abs, sheet_name):
    # global op_ips
    op_ips = []
    wb_obj = openpyxl.load_workbook(ip_xlsx_abs)
    sheet_obj = wb_obj.get_sheet_by_name(sheet_name)
    cols = PMIPCOLS
    print("max_rows", sheet_obj.max_row)
    print("max_columns", sheet_obj.max_column)

    for row in range(2, sheet_obj.max_row + 1):
        ip = sheet_obj.cell(row=row, column=cols.col_ip).value
        if ip is None:
            continue
        ip = ip.strip()
        host = sheet_obj.cell(row=row, column=cols.col_host).value.strip()
        so = sheet_obj.cell(row=row, column=cols.col_so).value.strip()
        user = sheet_obj.cell(row=row, column=cols.col_user).value
        pw = sheet_obj.cell(row=row, column=cols.col_pw).value
        if (user is None or len(user) == 0) or (pw is None or len(pw) == 0):
            user = device_user
            pw = device_pw
        else:
            user = user.strip()
            pw = pw.strip()
        op_ips.append([ip, host, so, user, pw])

    return op_ips


def citrix_cgnat_ssh_port_allocation_failure(w_dir, postfix_xtime, so_name, device_ip, device_user, device_pw,
                                             keyword):
    print("citrix_cgnat_ssh_port_allocation_failure", w_dir, postfix_xtime, so_name, device_ip, device_user, device_pw,
          keyword)
    log_dir = os.path.join(w_dir, postfix_xtime)
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    #    A. ssh to the ip
    ssh = paramiko.SSHClient()
    ssh.load_system_host_keys()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(device_ip,
                username=device_user,
                password=device_pw,
                look_for_keys=False)

    today = date.today()
    t_month = today.strftime("%b")
    t_day = today.strftime("%d")
    # t_day = str(int(t_day) - 1)
    cmd = "shell ls -lt /var/log/ns.log*"
    ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
    output = ssh_stdout.readlines()
    today_files = []
    # keyword = "portall"
    # keyword = "netScalerLoginFailure"
    h_template = 'shell {} -i "{}" {}'

    #    B. 取得檔案清單(程式執行當日) = log_files:
    for line in output:
        if "ns.log" in line:
            data = line.split()
            d_month = data[5]
            d_day = data[6]
            if d_month == t_month and d_day == t_day:
                print(line.strip())
                today_files.append(data[-1])
    #    C. 針對每一個log_file(device端):
    if len(today_files) > 0:
        lines_to_file = []
        for afile in today_files:
            print(afile, "----------------------------------------------------------------------------")
            grep = "zgrep" if afile.endswith(".gz") else "grep"
            cmd = h_template.format(grep, keyword, afile)
            ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
            output = ssh_stdout.readlines()
            for line in output:
                if line.strip().lower().startswith("warning") or line.strip().lower().startswith("done") or \
                        line.strip().lower().startswith("error"):
                    pass
                else:
                    lines_to_file.append(line)
                    print(line.rstrip())
        if len(lines_to_file) > 0:
            file_name = "{}_{}_{}.log".format(so_name, device_ip, postfix_xtime)
            file_abs = os.path.join(log_dir, file_name)
            print(file_abs)
            with open(file_abs, "w") as log_f_obj:
                log_f_obj.writelines(lines_to_file)
    ssh.close()


def citrix_cgnat_port_alloc_failure_grouping(log_dir_abs, xlsx_file_abs, grp_by_keyword):
    log_files = [f for f in os.listdir(log_dir_abs) if os.path.isfile(os.path.join(log_dir_abs,f)) and f.endswith(".log")]

    for log_file in log_files:
        f = open(os.path.join(log_dir, log_file), "r")
        data = f.readlines()
        f.close()
        flist = log_file.split("_")
        so_name = flist[0]
        ip = flist[1]
        so_dict = {}
        for idx, line in enumerate(data):
            print(idx, line.rstrip())
            if line.find(grp_by_keyword) > -1:
                dlist = line.split()
                kw_index = dlist.index(grp_by_keyword)
                if grp_by_keyword == "Remote_ip":
                    target_ip = dlist[kw_index+1]
                else:
                    target_ip = dlist[kw_index + 2][:-1]
                print(target_ip)
                if target_ip in so_dict.keys():
                    so_dict[target_ip][0] = so_dict[target_ip][0] + 1
                    so_dict[target_ip][1] = line.rstrip()
                else:
                    so_dict[target_ip] = [1, line.rstrip()]
        for xx in so_dict.items():
            print(xx[0], xx[1])


def process_command_line_args(args):
    global debug_mode
    global work_dir
    global device_user
    global device_pw
    print('Number of arguments:', len(args), 'arguments.')
    print('Argument List:', str(args))
    for arg_idx, arg in enumerate(args, start=0):
        print(arg_idx, arg)
        if arg == "-debug":
            debug_mode = True
        elif arg == "-wdir":
            work_dir = args[arg_idx+1]
        elif arg == "-u":
            device_user = args[arg_idx+1]
        elif arg == "-p":
            device_pw = args[arg_idx+1]


if __name__ == "__main__":
    process_command_line_args(sys.argv)
    postfix_xtime = datetime.datetime.now().strftime("%Y%m%d%H%M")
    if debug_mode:
        xlsx_ips = os.path.join(work_dir, "KBROXLSX/cgnat_ips_lab.xlsx")
        # postfix_xtime = "202103181419"
        # so_ips = [["idtlab01", "10.0.0.1"], ["idtlab02", "10.0.0.3"]]
        grep_keyword = "netScalerLoginFailure"
        grp_by_keyword = "Remote_ip"
    else:
        xlsx_ips = os.path.join(work_dir, "KBROXLSX/cgnat_ips.xlsx")
        grep_keyword = "portall"
        grp_by_keyword = "lsnSubscrIP"

    op_ips = xlsx_get_sites_and_telnet_infos(xlsx_ips, ip_sheet_name)
    log_dir = os.path.join(work_dir, postfix_xtime)
    xlsx_file_abs = os.path.join(work_dir, "portallfailed_{}.xlsx".format(postfix_xtime))
    for so_ip_info in op_ips:
        # [ip, host, so, user, pw]
        device_ip = so_ip_info[0]
        device_host = so_ip_info[1]
        so_name = so_ip_info[2]
        device_user = so_ip_info[3]
        device_pw = so_ip_info[4]
        citrix_cgnat_ssh_port_allocation_failure(work_dir, postfix_xtime, so_name, device_ip, device_user, device_pw,
                                                 grep_keyword)
    citrix_cgnat_port_alloc_failure_grouping(log_dir, xlsx_file_abs, grp_by_keyword)
