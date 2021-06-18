import datetime
import os
from datetime import date
import openpyxl

server_log_root_unc_path = r"\\10.0.100.90\73c20-區域\客戶資料"
server_asr = r"凱擘台媒ASR9K\MA PM 資料"
server_cgnat = "凱擘台媒CGNAT"
server_cmts = "凱擘台媒CMTS"

# M:\凱擘台媒CMTS\MA PM 資料\2021 凱 擘CMTS MA\Q1
kbro_so_with_index = ['01_UC', '02_KP', '03_HC', '04_NTP', '05_BNT', '06_DA_DAWS', '07_TC', '08_CG', '09_YMS', '10_Lab',
                      '12_YJL', '13_GVC', '14_UCT', '15_MCT', '16_FM', '17_NCC', '18_NT', '19_PHC', '20_PN', '21_RH',
                      '22_NTP_SI', '23_KS']

kbro_so_dict = {"全聯": '01_UC',           "金頻道": '02_KP',  "振道": '03_HC',   "新台北": '04_NTP', "北桃園": '05_BNT',
                "大安": '06_DA_DAWS',      "新唐城": '07_TC',  "大新店": '08_CG', "陽明山": '09_YMS', "LAB": '10_Lab',
                "永佳樂": '12_YJL',        "觀天下": '13_GVC', "聯禾": '14_UCT',  "紅樹林":'15_MCT',  "豐盟": '16_FM',
                "新頻道": '17_NCC',        "南天": '18_NT',    "鳳信": '19_PHC',  "屏南": '20_PN',    "瑞湖": '21_RH',
                "新台北信義": '22_NTP_SI', "觀昇": '23_KS'}

idt_domain = "idtech"
citrix_dont_write_cmds = ["vtysh", "ter le 0", "exit", "shell"]
n9k_dont_write_cmds = ["ter le 0", "ter len 0", "ter wi 511"]

pm_xlsx_file_name = ""
pm_user = "IDT_PM"
pm_pw = "IDT_PM@123"
cgnat_user = "citrix_pm"
cgnat_pw = "citrix_pm@123"
qb_user = "idt_tech"
qb_pw = "IDT_tech"
qb_viewer_user = "IDT_tech"
qb_viewer_pw = "IDT_tech"
qb_idrac_user = "idtech"
qb_idrac_pw = "Idtech123!"


quarter = (datetime.date.today().month-1)//3 + 1
year = datetime.date.today().year
str_now = datetime.datetime.now().strftime("%Y%m%d_%H%M")
# Modify20210511 作業存檔目錄由 LOG_ 改為 TEMP_
str_log = "TEMP_{}".format(str_now)
str_log_quarter = "LOG_{}_Q{}".format(year, quarter)


class IPCOLUMN:
    device_ip = 1
    device_host = 2
    device_so = 3
    device_type = 4
    device_user = 5
    device_pw = 6
    device_waittime = 7


class HomePlus_IPCOLUMN:
    device_ip = 1
    device_host = 2
    device_so = 3
    device_type = 4
    device_user = 5
    device_pw1 = 6
    device_pw2 = 7


def get_device_cmds_via_excel_file(device_type, kbro_pm_xlsx_file_name, str_op_quarter):
    device_cmds = []
    if not os.path.isfile(kbro_pm_xlsx_file_name):
        return device_cmds
    wb_obj = openpyxl.load_workbook(kbro_pm_xlsx_file_name)
    try:
        sheet_obj = wb_obj.get_sheet_by_name(device_type)
    except:
        print("No sheet in [{0}] named {1}".format(kbro_pm_xlsx_file_name, device_type))
        return device_cmds

    max_row = sheet_obj.max_row
    for row in range(2, max_row + 1):
        cmd = sheet_obj.cell(row=row, column=1).value
        # print(cmd)
        if cmd is None or len(cmd.strip()) == 0:
            pass
        else:
            cmd_quarter = sheet_obj.cell(row=row, column=2).value
            # print(cmd, cmd_quarter)
            if cmd_quarter is None or len(cmd_quarter.strip()) == 0:
                device_cmds.append(cmd.strip())
            else:
                if cmd_quarter.upper() == str_op_quarter.upper():
                    device_cmds.append(cmd.strip())
    wb_obj.close()
    return device_cmds


def get_ips_via_excel_file(kbro_pm_xlsx_file_name='KBRO PM.xlsx', sheet_name_ip="IP"):
    global pm_xlsx_file_name
    pm_xlsx_file_name = kbro_pm_xlsx_file_name
    if not os.path.isfile(kbro_pm_xlsx_file_name):
        return False
    wb_obj = openpyxl.load_workbook(kbro_pm_xlsx_file_name)
    try:
        sheet_obj = wb_obj.get_sheet_by_name(sheet_name_ip)
    except:
        print("No sheet in [{0}] named {1}".format(kbro_pm_xlsx_file_name, sheet_name_ip))
        return False

    cols = IPCOLUMN()
    pm_ip_list = []
    # today = date.today()
    # print("Today's date:", today, today.month, today.day)
    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column
    print("get_ips_via_excel_file: (max_row, max_column)", max_row, max_column)
    for row in range(2, max_row + 1):
        device_ip = sheet_obj.cell(row=row, column=cols.device_ip).value
        if device_ip is None or len(device_ip.strip()) == 0:
            continue
        else:
            device_ip = device_ip.strip()
            device_host = sheet_obj.cell(row=row, column=cols.device_host).value
            device_host = "無名" if device_host is None else device_host.strip()
            device_so = sheet_obj.cell(row=row, column=cols.device_so).value.strip()
            device_type = sheet_obj.cell(row=row, column=cols.device_type).value.strip()
            device_user = sheet_obj.cell(row=row, column=cols.device_user).value
            device_pw = sheet_obj.cell(row=row, column=cols.device_pw).value
            if device_user is None or len(device_user.strip()) == 0:
                if device_type.upper() == "QB":
                    device_user = qb_user
                    device_pw = qb_pw
                elif device_type.upper() == "CGNAT":
                    device_user = cgnat_user
                    device_pw = cgnat_pw
                else:
                    device_user = pm_user
                    device_pw = pm_pw
            else:
                device_user = device_user.strip()
                if device_pw is None or len(device_pw.strip()) == 0:
                    device_pw = ""
            # Modify20210618 新增針對網頁截圖特別指定的等待時間
            device_waittime = sheet_obj.cell(row=row, column=cols.device_waittime).value
            if device_waittime is None:
                device_waittime = -1

            pm_ip_list.append([device_ip, device_host, device_so, device_type, device_user, device_pw, device_waittime])
    wb_obj.close()
    return pm_ip_list


def get_ips_via_excel_homeplus(the_pm_xlsx_file_name, sheet_name_ip="IP"):
    global pm_xlsx_file_name
    pm_xlsx_file_name =the_pm_xlsx_file_name
    if not os.path.isfile(the_pm_xlsx_file_name):
        return False
    wb_obj = openpyxl.load_workbook(the_pm_xlsx_file_name)
    try:
        sheet_obj = wb_obj.get_sheet_by_name(sheet_name_ip)
    except:
        print("No sheet in [{0}] named {1}".format(the_pm_xlsx_file_name, sheet_name_ip))
        return False

    cols = HomePlus_IPCOLUMN()
    pm_ip_list = []
    # today = date.today()
    # print("Today's date:", today, today.month, today.day)
    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column
    print("get_ips_via_excel_file: (max_row, max_column)", max_row, max_column)
    for row in range(2, max_row + 1):
        device_ip = sheet_obj.cell(row=row, column=cols.device_ip).value
        if device_ip is None or len(device_ip.strip()) == 0:
            continue
        else:
            device_ip = device_ip.strip()
            device_host = sheet_obj.cell(row=row, column=cols.device_host).value
            device_host = "無名" if device_host is None else device_host.strip()
            device_so = sheet_obj.cell(row=row, column=cols.device_so).value.strip()
            device_type = sheet_obj.cell(row=row, column=cols.device_type).value.strip()
            device_user = sheet_obj.cell(row=row, column=cols.device_user).value
            device_pw1 = sheet_obj.cell(row=row, column=cols.device_pw1).value
            device_pw2 = sheet_obj.cell(row=row, column=cols.device_pw2).value
            # todo 對中嘉(沒有TACAS)而言: device_user=第一層密碼, device_pw=第二層密碼
            if device_user is None or len(device_user.strip()) == 0:
                device_user = ""
                device_pw1 = ""
                device_pw2 = ""
            else:
                device_user = device_user.strip()
                if device_pw1 is None or len(device_pw1.strip()) == 0:
                    device_pw1 = ""
                if device_pw2 is None or len(device_pw2.strip()) == 0:
                    device_pw2 = ""
            pm_ip_list.append([device_ip, device_host, device_so, device_type, device_user, device_pw1, device_pw2])
    wb_obj.close()
    return pm_ip_list