import os
import xlwings as xw
import openpyxl
import idt_tools_xlsx as idtxlsx


def cgnat_port_count_write_header(sheet, row_index):
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, 1, "Dest Port", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, 2, "PortSessionCount", border_type=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, 3, "%", border_type=1, align=1)


def cgnat_device_port_count_to_xlsx(sheet, dest_port_count_dict, total):
    # port_and_counts.sort(key=lambda x: x[1], reverse=True)
    number_of_data = len(dest_port_count_dict)
    write_percent_count = 20
    row_index = 1
    cgnat_port_count_write_header(sheet, row_index)
    for idx, key in enumerate(dest_port_count_dict.keys()):
        row_index += 1
        bt = 1 if idx < write_percent_count else 0
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, 1, key,
                                                     border_type=bt, align=1)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, 2, dest_port_count_dict[key],
                                                     border_type=bt, align=1)
        if idx < write_percent_count:
            percent = "{0:.2%}".format(dest_port_count_dict[key]/total)
            idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, 3, percent, border_type=1, align=1)
    print(total)


def init_bucket(keys):
    init_dict = {}
    for key in keys:
        init_dict[key] = 0
    return init_dict


def add_bucket_count(value, range_keys, range_dict):
    for idx, key in enumerate(range_keys):
        if idx == 0:
            if value <= key:
                range_dict[key] += 1
                break
        else:
            if range_keys[idx - 1] < value <= key:
                range_dict[key] += 1
                break


def cgnat_session_count_write_header(sheet, row_index, start_col):
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, "級距", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col+1, "IP數量", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col+2, "%", border_type=1, align=1)


def cgnat_session_count_write_data(sheet, row_index, range_keys, range_dict, sum_all, start_col):
    row_start = row_index
    for xx in range_keys:
        row_index += 1
        print(xx, range_dict[xx], "{0:.2%}".format(range_dict[xx]/sum_all))
        # sheet.cell(row=row_index, column=start_col).value = xx
        # sheet.cell(row=row_index, column=start_col+1).value = range_dict[xx]
        percent = "{0:.2%}".format(range_dict[xx]/sum_all)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, xx, border_type=1, align=1)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, range_dict[xx], border_type=1, align=1)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 2, percent, border_type=1, align=1)

    row_index += 1
    dict_sum = sum(range_dict.values())
    percent = "{0:.2%}".format(dict_sum/sum_all)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, "Total", align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, dict_sum, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 2, percent, align=1)


def cgnat_device_session_count_to_xlsx(ip_session_dict, sheet, total):
    ip_keys_all = [1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000, 20000, 30000, 40000, 50000, 60000,
                   70000, 10000]
    ip_keys_1000 = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
    ip_keys_200 = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120, 130, 140, 150, 160, 170, 180, 190, 200]
    dict_all = init_bucket(ip_keys_all)
    dict_1000 = init_bucket(ip_keys_1000)
    dict_200 = init_bucket(ip_keys_200)

    row_index = 1
    start_col = 1
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, "IP", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col+1, "Session Count", border_type=1, align=1)
    for ip in ip_session_dict.keys():
        row_index += 1
        ip_session_count = ip_session_dict[ip]
        if start_col == 1:
            idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, ip, align=2)
            idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, ip_session_count, align=1)
        if row_index > 1:
            if ip_session_count <= ip_keys_200[-1]:
                add_bucket_count(ip_session_count, ip_keys_200, dict_200)
            if ip_session_count <= ip_keys_1000[-1]:
                add_bucket_count(ip_session_count, ip_keys_1000, dict_1000)
            add_bucket_count(ip_session_count, ip_keys_all, dict_all)

    sum_all = sum(dict_all.values())
    row_index = 1
    if start_col == 1:
        col_start_all = start_col + 4
    else:
        col_start_all = start_col
    cgnat_session_count_write_header(sheet, row_index, col_start_all)
    cgnat_session_count_write_data(sheet, row_index, ip_keys_all, dict_all, sum_all, col_start_all)

    row_index = 1
    col_start_1000 = col_start_all + 4
    cgnat_session_count_write_header(sheet, row_index, col_start_1000)
    cgnat_session_count_write_data(sheet, row_index, ip_keys_1000, dict_1000, sum_all, col_start_1000)

    row_index = 1
    col_start_200 = col_start_1000 + 4
    cgnat_session_count_write_header(sheet, row_index, col_start_200)
    cgnat_session_count_write_data(sheet, row_index, ip_keys_200, dict_200, sum_all, col_start_200)
    print(sum_all)


def cgnat_so_device_log_to_xlsx(device_log_abs, so_as_key, device_host, xtime):
    device_xlsx_abs = os.path.join(os.path.dirname(device_log_abs), f"{xtime}_{so_as_key}_{device_host}.xlsx")
    # device_xlsx_abs = device_log_abs[:-4] + ".xlsx"
    sn_port = f"{so_as_key}{device_host[-1]}_PORT"
    sn_ip = f"{so_as_key}{device_host[-1]}_IP"
    dest_port_dict = {}
    source_ip_dict = {}
    data_count = 0
    with open(device_log_abs, "r") as file_obj:
        while True:
            line = file_obj.readline()
            if not line:
                break
            data_list = line.strip().split()
            if 10 < len(data_list):
                data_count += 1
                src_ip = data_list[1]
                dest_port = data_list[5]
                if src_ip in source_ip_dict.keys():
                    source_ip_dict[src_ip] = source_ip_dict[src_ip] + 1
                else:
                    source_ip_dict[src_ip] = 1

                if dest_port in dest_port_dict.keys():
                    dest_port_dict[dest_port] = dest_port_dict[dest_port] + 1
                else:
                    dest_port_dict[dest_port] = 1

    sorted_src_ip = {k: v for k, v in sorted(source_ip_dict.items(), key=lambda item: item[1], reverse=True)}
    for idx, src_ip in enumerate(sorted_src_ip.keys()):
        if 10 < idx:
            break
        print("sorted_src_ip:", idx, src_ip, sorted_src_ip[src_ip])

    sorted_dest_port = {k: v for k, v in sorted(dest_port_dict.items(), key=lambda item: item[1], reverse=True)}
    for idx, src_ip in enumerate(sorted_dest_port.keys()):
        if 10 < idx:
            break
        print("sorted_dest_port:", idx, src_ip, sorted_dest_port[src_ip], sorted_dest_port[src_ip]*100/data_count)
    # todo dictionary to xlsx
    #  1. 頁簽 [SO]_DEST_PORT
    #  2. 頁簽 [SO]_SRC_IP
    wb = openpyxl.Workbook()
    ws_dest_port = wb.create_sheet(sn_port)
    cgnat_port_count_write_header(ws_dest_port, 1)
    cgnat_device_port_count_to_xlsx(ws_dest_port, sorted_dest_port, data_count)
    ws_src_ip = wb.create_sheet(sn_ip)
    cgnat_device_session_count_to_xlsx(sorted_src_ip, ws_src_ip, data_count)
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove_sheet(default_sheet)
    wb.save(device_xlsx_abs)


def cgnat_merge_xlsx(data_dir, xminute):
    xlsx_abs = os.path.join(data_dir, f"CGNAT統計_{xminute}.xlsx")
    xlsxs_to_merge = [os.path.join(data_dir, f) for f in os.listdir(data_dir) if f.startswith(xminute) and f.endswith(".xlsx")]
    wb_new = xw.Book()
    app_new = wb_new.app
    for xlsx in xlsxs_to_merge:
        wb_orig = xw.Book(xlsx)
        for sheet in wb_orig.sheets:
            sheet_name = sheet.name
            print("sheet to copy:", sheet_name)
            sheet.api.Copy(Before=wb_new.sheets[0].api)
        wb_orig.close()
    wb_new.save(xlsx_abs)
    wb_new.close()
    app_new.quit()
    return xlsx_abs


