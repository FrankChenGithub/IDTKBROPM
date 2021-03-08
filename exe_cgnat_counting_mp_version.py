import datetime
import sys

import openpyxl
import psutil
from openpyxl.styles.borders import Border, Side
import os
from os.path import isfile, isdir, join
from os import listdir
import idt_tools_xlsx as idtxlsx
from multiprocessing import Process, freeze_support


def exe_parselog_pl(command):
    os.system(command)


def find_folder_list(work_dir):
    xlsx_list = []
    op_folder = []
    xlsx_files = [f[:-5] for f in listdir(work_dir) if f[-4:] == "xlsx" and isfile(join(work_dir, f))]
    print("xlsx_files", xlsx_files)
    xlsx_files = sorted(xlsx_files)
    for idx, xlsx_file in enumerate(xlsx_files, start=0):
        xlsx_list.append(xlsx_file)
    subdirs = [f for f in os.listdir(work_dir) if os.path.isdir(os.path.join(work_dir, f))]
    if len(xlsx_files) > 0:
        # remove some folder from the list
        for opdir in subdirs:
            need_op = True
            for xlsx_file in xlsx_files:
                YYYYMMDD = xlsx_file.split("_")[-1]
                if opdir == YYYYMMDD:
                    need_op = False
                    break
            if need_op:
                op_folder.append(opdir)
    else:
        op_folder = subdirs
    return op_folder


def init_bucket(keys):
    init_dict = {}
    for key in keys:
        init_dict[key] = 0
    return init_dict


def write_port_count_header(sheet, row_index, start_col):
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, "DistPort", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, "PortSessionCount", border_type=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 2, "%", border_type=1, align=1)


def exe_csv_port_count(csv_file, sheet, start_col=1):
    count = 0
    port_and_counts = []
    total = 0
    with open(csv_file, mode="r") as csv_lines:
        for csv in csv_lines:
            count += 1
            if count > 1:
                pc = csv.strip().split(",")
                port_count =[int(pc[0]), int(pc[1])]
                port_and_counts.append(port_count)
                total += int(pc[1])
    port_and_counts.sort(key=lambda x: x[1], reverse=True)
    number_of_row = len(port_and_counts)
    write_percent_count = 20
    row_index = 1
    write_port_count_header(sheet, row_index, start_col)
    write_out_data_count = number_of_row if start_col == 1 else write_percent_count
    for idx in range(write_out_data_count):
        row_index += 1
        bt = 1 if idx < write_percent_count else 0
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, port_and_counts[idx][0],
                                                     border_type=bt, align=1)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, port_and_counts[idx][1],
                                                     border_type=bt, align=1)
        if idx < write_percent_count:
            percent = "{0:.2%}".format(port_and_counts[idx][1]/total)
            # print(port_and_counts[idx][0], port_and_counts[idx][1], percent)
            idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 2, percent, border_type=1, align=1)
    print(total)


def add_bucket_count(value, range_keys, range_dict):
    for idx, key in enumerate(range_keys):
        if idx == 0:
            if value <= key:
                range_dict[key] += 1
                break
        else:
            if range_keys[idx - 1] < value <= key:
                range_dict[key] += 1
                break


def write_session_count_header(sheet, row_index, start_col):
    # sheet.cell(row=row_index, column=start_col).value = "級距"
    # sheet.cell(row=row_index, column=start_col + 1).value = "IP數量"
    # sheet.cell(row=row_index, column=start_col + 2).value = "%"
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, "級距", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col+1, "IP數量", border_type=1, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col+2, "%", border_type=1, align=1)


def write_session_count_data(sheet, row_index, range_keys, range_dict, sum_all, start_col):
    row_start = row_index
    for xx in range_keys:
        row_index += 1
        print(xx, range_dict[xx], "{0:.2%}".format(range_dict[xx]/sum_all))
        # sheet.cell(row=row_index, column=start_col).value = xx
        # sheet.cell(row=row_index, column=start_col+1).value = range_dict[xx]
        percent = "{0:.2%}".format(range_dict[xx]/sum_all)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, xx, border_type=1, align=1)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, range_dict[xx], border_type=1, align=1)
        idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 2, percent, border_type=1, align=1)

    row_index += 1
    dict_sum = sum(range_dict.values())
    percent = "{0:.2%}".format(dict_sum/sum_all)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, "Total", align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, dict_sum, align=1)
    idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 2, percent, align=1)


def exe_csv_session_count(csv_file, sheet, start_col=1):
    ip_keys_all = [1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000, 20000, 30000, 40000, 50000, 60000,
                   70000, 10000]
    ip_keys_1000 = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
    ip_keys_200 = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110, 120, 130, 140, 150, 160, 170, 180, 190, 200]
    dict_all = init_bucket(ip_keys_all)
    dict_1000 = init_bucket(ip_keys_1000)
    dict_200 = init_bucket(ip_keys_200)
    # count = 0
    row_index = 0
    with open(csv_file, mode="r") as csv_lines:
        for csv in csv_lines:
            row_index += 1
            ip_and_count = csv.strip().split(",")
            if start_col == 1:
                idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col, ip_and_count[0], align=2)
                idtxlsx.write_cell_with_border_and_alignment(sheet, row_index, start_col + 1, ip_and_count[1], align=1)
            if row_index > 1:
                ip_sess_count = int(ip_and_count[1])
                if ip_sess_count <= ip_keys_200[-1]:
                    add_bucket_count(ip_sess_count, ip_keys_200, dict_200)
                if ip_sess_count <= ip_keys_1000[-1]:
                    add_bucket_count(ip_sess_count, ip_keys_1000, dict_1000)
                add_bucket_count(ip_sess_count, ip_keys_all, dict_all)

    sum_all = sum(dict_all.values())
    row_index = 1
    if start_col == 1:
        col_start_all = start_col + 4
    else:
        col_start_all = start_col
    write_session_count_header(sheet, row_index, col_start_all)
    write_session_count_data(sheet, row_index, ip_keys_all, dict_all, sum_all, col_start_all)

    row_index = 1
    col_start_1000 = col_start_all + 4
    write_session_count_header(sheet, row_index, col_start_1000)
    write_session_count_data(sheet, row_index, ip_keys_1000, dict_1000, sum_all, col_start_1000)

    row_index = 1
    col_start_200 = col_start_1000 + 4
    write_session_count_header(sheet, row_index, col_start_200)
    write_session_count_data(sheet, row_index, ip_keys_200, dict_200, sum_all, col_start_200)
    print(sum_all)


def cgnat_mp_logparsing_physical_cpu(wdir, log_dir, ide_mode):
    physical_cpu = psutil.cpu_count(logical=False)
    time_start = datetime.datetime.now()
    # parse_pl = os.path.join(wdir, "ParsingLog.exe")
    parse_pl = join(wdir, "ParsingLog.{}".format("pl" if ide_mode else "exe"))
    log_dir_abs = os.path.join(wdir, log_dir)
    print(wdir, log_dir_abs)
    log_files = [f for f in os.listdir(log_dir_abs) if f[-3:] == "log"]
    log_files_count = len(log_files)
    for log_idx in range(0, log_files_count, physical_cpu):
        log_procs = []
        log_step_end = log_idx+physical_cpu if log_idx+physical_cpu < log_files_count else log_files_count
        for idx in range(log_idx, log_step_end):
            log_file = log_files[idx]
            so_name = log_file[:log_file.rfind("_")].split("_")[-1]
            lop_file_abs = os.path.join(log_dir_abs, log_file)
            command = "{} {}".format(parse_pl, lop_file_abs)
            print(idx, so_name, datetime.datetime.now(), command)
            proc = Process(target=exe_parselog_pl, args=(command,))
            log_procs.append(proc)
            proc.start()
        for proc in log_procs:
            proc.join()
    time_end = datetime.datetime.now()
    print("start@", time_start)
    print("end@", time_end)
    print("total:", time_end - time_start)


def cgnat_mp_logparsing_ide(wdir, log_dir, ide_mode):
    time_start = datetime.datetime.now()
    # parse_pl = os.path.join(wdir, "ParsingLog.exe")
    parse_pl = join(wdir, "ParsingLog.{}".format("pl" if ide_mode else "exe"))
    log_dir_abs = os.path.join(wdir, log_dir)
    print(wdir, log_dir_abs)

    log_files = [f for f in os.listdir(log_dir_abs) if f[-3:] == "log"]
    log_procs = []
    for idx, log_file in enumerate(log_files):
        so_name = log_file[:log_file.rfind("_")].split("_")[-1]
        lop_file_abs = os.path.join(log_dir_abs, log_file)
        command = "{} {}".format(parse_pl, lop_file_abs)
        print(idx, so_name, datetime.datetime.now(), command)
        # os.system(command)
        proc = Process(target=exe_parselog_pl, args=(command,))
        log_procs.append(proc)
        proc.start()
        # proc.join()
        # print(idx, datetime.datetime.now(), command)

    for proc in log_procs:
        proc.join()
    time_end = datetime.datetime.now()
    print("start@", time_start)
    print("end@", time_end)
    print("total:", time_end - time_start)


def cgnat_csvs_to_xlsx(wdir, csv_sub):
    xlsx_nt = "CGNAT統計_{}.xlsx"
    time_start = datetime.datetime.now()
    xlsx_file = join(wdir, xlsx_nt.format(csv_sub))
    wb = openpyxl.Workbook()
    print(xlsx_file)
    csv_dir_abs = join(wdir, csv_sub)
    print(csv_dir_abs)
    csv_files = [f for f in listdir(csv_dir_abs) if f[-3:] == "csv"]
    for idx, csv_file in enumerate(csv_files):
        att_list = csv_file.split("_")
        so_name = att_list[2]
        port_or_ip = "_PORT" if att_list[-1] == "DstPortSessionCount.csv" else "_IP"
        csv_file_abs = os.path.join(csv_dir_abs, csv_file)
        ws_name = so_name + port_or_ip
        print(idx, ws_name, datetime.datetime.now())
        ws = wb.create_sheet(ws_name)
        if port_or_ip == "_PORT":
            exe_csv_port_count(csv_file_abs, ws)
        else:
            exe_csv_session_count(csv_file_abs, ws)
        idtxlsx.auto_adjust_column_width(ws)
    print(csv_sub, "Before Save @", datetime.datetime.now())
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(xlsx_file)
    time_end = datetime.datetime.now()
    print("start@", time_start)
    print("end@", time_end)
    print("total:", time_end - time_start)


def cgnat_main_multiprocessing(wdir, ide_mode):
    xlsx_nt = "CGNAT統計_{}.xlsx"
    time_start = datetime.datetime.now()
    parse_pl = join(wdir, "ParsingLog.{}".format("pl" if ide_mode else "exe"))
    print(wdir)
    op_folders = find_folder_list(wdir)
    print(op_folders)
    for log_dir in op_folders:
        cgnat_mp_logparsing_ide(wdir, log_dir, ide_mode)
        # cgnat_mp_logparsing_physical_cpu(wdir, log_dir)

    for op_folder in op_folders:
        cgnat_csvs_to_xlsx(wdir, op_folder)
        # xlsx_file = join(wdir, xlsx_nt.format(op_folder))
        # wb = openpyxl.Workbook()
        # print(xlsx_file)
        # op_dir_full = join(wdir, op_folder)
        # print(op_dir_full)
        # op_files = [f for f in listdir(op_dir_full) if f[-3:] == "log"]
        # for idx, op_file in enumerate(op_files):
        #     so_name = op_file[:op_file.rfind("_")].split("_")[-1]
        #     op_file = join(op_dir_full, op_file)
        #     print(idx, so_name , datetime.datetime.now())
        #     port_count_csv = "{}_DstPortSessionCount.csv".format(op_file[:-4])
        #     ws_port = wb.create_sheet(so_name + "_PORT")
        #     exe_csv_port_count(port_count_csv, ws_port)
        #     idtxlsx.auto_adjust_column_width(ws_port)
        #     ip_sess_csv = "{}_SrcIpSessionCount.csv".format(op_file[:-4])
        #     ws_ip = wb.create_sheet(so_name + "_IP")
        #     exe_csv_session_count(ip_sess_csv, ws_ip)
        #     idtxlsx.auto_adjust_column_width(ws_ip)
        # print(op_folder, "Before Save @", datetime.datetime.now())
        # if "Sheet" in wb.sheetnames:
        #     del wb["Sheet"]
        # wb.save(xlsx_file)
        time_end = datetime.datetime.now()
        print("start@", time_start)
        print("end@", time_end)
        print("total:", time_end - time_start)


if __name__ == "__main__":
    freeze_support()
    ide_mode = False
    if len(sys.argv) > 1:
        if sys.argv[0].lower().endswith(".exe"):
            if sys.argv[1].lower().endswith(".py"):
                ide_mode = True
    # root_wdir = os.getcwd()
    root_wdir = r"C:\cgnat"
    xlsx_nt = "CGNAT統計_{}.xlsx"
    time_start = datetime.datetime.now()
    cgnat_main_multiprocessing(root_wdir, ide_mode)
    time_end = datetime.datetime.now()
    print("__main__ start@", time_start)
    print("__main__ end@", time_end)
    print("__main__  total:", time_end - time_start)

